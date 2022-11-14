import requests
import json
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import tzlocal

def obtain_access_token():
    url = "https://www.fflogs.com/oauth/token"

    payload={'grant_type': 'client_credentials'}
    file = open("C:\\Users\\AbsoluteCrow\\Documents\\Programming\\Python\\fflogs_keys.json")
    fflogs_keys = json.load(file)

    headers = {
    'Authorization': f"{fflogs_keys['Auth_Basic']}"
    }

    response = requests.request("POST", url, headers=headers, data=payload).json()
    return response['access_token']

def query_fflogs(graphql_query):
    url = "https://www.fflogs.com/api/v2/client"

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    data = {"query": graphql_query}
    json_data = json.dumps(data)

    response = requests.request("POST", url, headers=headers, data=json_data).json()
    return response

def search_through_deaths_for_name(deaths, name):
    death_counter = 0
    for death in deaths['data']['reportData']['report']['table']['data']['entries']:
        if (death['name'] == name):
            death_counter += 1
    return death_counter

# Set Excel Sheet Variables
logs_excel = 'P:/My Drive/DSR - Logs.xlsx'
workbook = load_workbook(logs_excel)
sheet = workbook.worksheets[0]
COL_LOGS = 'B'
COL_DATE = 'C'
COL_WIPES = 'A'
COL_CES = 'G'
COL_JAMIE = 'H'
COL_FAYTH = 'I'
COL_EVE = 'J'
COL_KAEDE = 'K'
COL_MITHRIL = 'L'
COL_FERESIA = 'M'
COL_RALF = 'N'
COL_AVERAGE_FIGHT_PERCENTAGE = 'D'
COL_BEST_FIGHT_PERCENTAGE = 'E'
COL_PHASE1 = 'O'
COL_PHASE2 = 'P'
COL_PHASE3 = 'Q'
COL_PHASE4 = 'R'
COL_PHASE_INT = 'S'
COL_PHASE5 = 'T'
COL_PHASE6 = 'U'

access_token = obtain_access_token()

# Query API for the Codes Listed
for i in range(3,sheet.max_row):
    current_location = f"{COL_LOGS}{i}"

    if ((sheet[current_location].value is not None) and (sheet[f"{COL_AVERAGE_FIGHT_PERCENTAGE}{i}"].value is None)):

        # Iterate through the excel sheet url to obtain the URL of reach report.
        code = re.search(r"reports\/(.+)\/", sheet[current_location].value).group(1)
        graphql_query = f"""
        {{
            reportData{{
                report(code: "{code}"){{
                    startTime
                    fights(encounterID: 1065){{
                        id
                        endTime
                        fightPercentage
                        lastPhase
                    }}
                }}
            }}
        }}
        """
        print(f"Processing {code}")

        try:
            fights = query_fflogs(graphql_query) # Query against the FF Logs api to obtain the fights in the URL
        except:
            fights = query_fflogs(graphql_query)

        # Extract and Store the date in the Excel Sheet column C
        local_timezone = tzlocal.get_localzone()
        unix_timestamp = float(fights['data']['reportData']['report']['startTime'])/1000
        local_time = datetime.fromtimestamp(unix_timestamp, local_timezone)
        sheet[f"{COL_DATE}{i}"] = local_time.strftime(r"%m/%d/%Y")

        # Preparing the Excel before processing fights.
        sheet[f"{COL_CES}{i}"] = int(0)
        sheet[f"{COL_JAMIE}{i}"] = int(0)
        sheet[f"{COL_FAYTH}{i}"] = int(0)
        sheet[f"{COL_EVE}{i}"] = int(0)
        sheet[f"{COL_KAEDE}{i}"] = int(0)
        sheet[f"{COL_MITHRIL}{i}"] = int(0)
        sheet[f"{COL_FERESIA}{i}"] = int(0)
        sheet[f"{COL_RALF}{i}"] = int(0)

        # Iterate through each fight in the report and add up the deaths.
        best_percentage = 100.0
        fight_counter = 0
        sum = 0.00
        p1_wipe_counter = 0
        p2_wipe_counter = 0
        p3_wipe_counter = 0
        p4_wipe_counter = 0
        int_wipe_counter = 0
        p5_wipe_counter = 0
        p6_wipe_counter = 0

        for fight in fights['data']['reportData']['report']['fights']:
            fight_counter += 1
            sum += fight['fightPercentage']

            if (float(fight['fightPercentage']) < best_percentage):
                best_percentage = float(fight['fightPercentage'])

            match int(fight['lastPhase']):
                case 1:
                    if (float(fight['fightPercentage']) > 80.0):
                        p1_wipe_counter += 1
                    elif (float(fight['fightPercentage']) < 80.0):
                        int_wipe_counter += 1
                case 2:
                    p2_wipe_counter += 1
                case 3:
                    p3_wipe_counter += 1
                case 4:
                    p4_wipe_counter += 1
                case 5:
                    p5_wipe_counter += 1
                case 6:
                    p6_wipe_counter += 1

            graphql_query = f"""
            {{
                reportData{{
                    report(code: "{code}"){{
                        table(fightIDs:{fight['id']}, endTime:{fight['endTime']}, dataType:Deaths)
                    }}
                }}
            }}
            """
            try:
                deaths = query_fflogs(graphql_query)
            except:
                deaths = query_fflogs(graphql_query)
            
            death_counter = search_through_deaths_for_name(deaths, "Cesyanis Corvus")
            sheet[f"{COL_CES}{i}"] = int(death_counter) + int(sheet[f"{COL_CES}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Jamie Rosefall")
            sheet[f"{COL_JAMIE}{i}"] = int(death_counter) + int(sheet[f"{COL_JAMIE}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Fayth Delarosa")
            sheet[f"{COL_FAYTH}{i}"] = int(death_counter) + int(sheet[f"{COL_FAYTH}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Evlesk Soimer")
            sheet[f"{COL_EVE}{i}"] = int(death_counter) + int(sheet[f"{COL_EVE}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Kaede Yasashi")
            sheet[f"{COL_KAEDE}{i}"] = int(death_counter) + int(sheet[f"{COL_KAEDE}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Mithril Dieties")
            sheet[f"{COL_MITHRIL}{i}"] = int(death_counter) + int(sheet[f"{COL_MITHRIL}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Feresia Peith")
            sheet[f"{COL_FERESIA}{i}"] = int(death_counter) + int(sheet[f"{COL_FERESIA}{i}"].value)

            death_counter = search_through_deaths_for_name(deaths, "Ralphonso Deluxe")
            sheet[f"{COL_RALF}{i}"] = int(death_counter) + int(sheet[f"{COL_RALF}{i}"].value)

        sheet[f"{COL_WIPES}{i}"] = int(fight_counter)
        sheet[f"{COL_AVERAGE_FIGHT_PERCENTAGE}{i}"] = float(sum/fight_counter)
        sheet[f"{COL_BEST_FIGHT_PERCENTAGE}{i}"] = float(best_percentage)
        sheet[f"{COL_PHASE1}{i}"] = int(p1_wipe_counter)
        sheet[f"{COL_PHASE2}{i}"] = int(p2_wipe_counter)
        sheet[f"{COL_PHASE3}{i}"] = int(p3_wipe_counter)
        sheet[f"{COL_PHASE4}{i}"] = int(p4_wipe_counter)
        sheet[f"{COL_PHASE_INT}{i}"] = int(int_wipe_counter)
        sheet[f"{COL_PHASE5}{i}"] = int(p5_wipe_counter)
        sheet[f"{COL_PHASE6}{i}"] = int(p6_wipe_counter)

workbook.save(logs_excel)